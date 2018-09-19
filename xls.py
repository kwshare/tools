# coding: utf-8
# WebTerminal2 - xls.py
# 2018/9/13
# package required: openpyxl xlrd xlwt xlutils


import itertools
import string
import re

import xlrd
import xlwt

from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
from xlutils.filter import process, XLRDReader, XLWTWriter


class InsertColumn:
    def __init__(self, read_file, write_file):
        if 'xlsx' not in read_file:
            self.__old = True

            self.__write_file = write_file

            self.read_wb = xlrd.open_workbook('test.xls', formatting_info=True)
            self.read_ws = self.read_wb.sheet_by_index(0)

            self.write_wb = xlwt.Workbook()
            self.write_ws = self.write_wb.add_sheet(self.read_ws.name, cell_overwrite_ok=True)

            self.__nrows = self.read_ws.nrows
            self.__ncols = self.read_ws.ncols
            self.__outStyle = self.__copy_style(self.read_wb)
        else:
            self.__old = False

            self.__write_file = write_file

            self.read_wb = load_workbook(read_file)
            self.write_wb = Workbook()

            self.read_ws = self.read_wb.active
            self.write_ws = self.write_wb.active

            self.write_ws.title = self.read_ws.title

            self.__al = Alignment(horizontal="center", vertical="center")
            __thin = Side(border_style="thin", color="000000")
            self.__border = Border(top=__thin, left=__thin, right=__thin, bottom=__thin)
            self.__table = list(itertools.chain(string.ascii_uppercase,
                                                (''.join(pair) for pair in
                                                 itertools.product(string.ascii_uppercase, repeat=2))))

    def __del__(self):
        if not self.__old:
            self.read_wb.close()
            self.write_wb.close()

    def __generate_style(self, x, y):
        xf_index = self.read_ws.cell_xf_index(x, y)
        saved_style = self.__outStyle[xf_index]
        return saved_style

    @staticmethod
    def __copy_style(wb):
        w = XLWTWriter()
        process(XLRDReader(wb, 'unknown.xls'), w)
        return w.style_list

    def insert_col(self, size):
        if self.__old:
            self.__insert_col_old(size)
        else:
            self.__insert_col_new(size)

    def __insert_col_old(self, size):

        for (rlow, rhigh, clow, chigh) in self.read_ws.merged_cells:
            # 第二个和第四个参数要减1
            self.write_ws.write_merge(rlow, rhigh - 1, clow + size,
                                      chigh - 1 + size, self.read_ws.cell(rlow, clow).value)

        for i in range(self.__nrows):
            for j in range(self.__ncols):
                # 尝试着能否在这里复制样式
                if self.read_ws.cell(i, j).value:
                    self.write_ws.write(i, j + size, self.read_ws.cell(i, j).value,
                                        style=self.__generate_style(i, j))
                elif isinstance(self.read_ws.cell(i, j).value, float):
                    self.write_ws.write(i, j + size, self.read_ws.cell(i, j).value)
                else:
                    try:
                        self.write_ws.write(i, j + size, self.read_ws.cell(i, j).value,
                                            style=self.__generate_style(i, j))
                    except ValueError:
                        pass

    def __insert_col_new(self, size):
        for i in self.read_ws.rows:
            for v in i:
                coord = '%s%s' % (self.__twenty_six(self.__get_alpha(v.coordinate), size),
                                  self.__get_int(v.coordinate))
                # fill border and everything
                self.write_ws[coord].fill = copy(v.fill)
                self.write_ws[coord].font = copy(v.font)
                self.write_ws[coord].border = copy(v.border)
                if v.value:
                    self.write_ws[coord] = v.value
                elif isinstance(v.value, long):
                    self.write_ws[coord] = v.value

        # applying styles
        for merge in self.read_ws.merged_cells.ranges:
            self.__style_range(self.write_ws, self.__move_right(merge.coord, size), border=self.__border,
                               alignment=self.__al)

    @staticmethod
    def __get_alpha(s):
        result = ''.join(re.split(r'[^A-Z]', s))
        return result

    @staticmethod
    def __get_int(s):
        return re.findall("\d+", s)[0]

    def __twenty_six(self, char, size):
        # 把坐标移动size个单位
        # list vs [{k:v},{v:k}] almost the same performance under 676 items.
        index = self.__table.index(char)
        return self.__table[index + size]

    def __move_right(self, coord, size):
        # A1:C2
        # 左半部分 右半部分
        left = coord.split(':')[0]
        right = coord.split(':')[1]

        left_alpha = self.__get_alpha(left)
        left_int = self.__get_int(left)
        right_alpha = self.__get_alpha(right)
        right_int = self.__get_int(right)

        return "%s%s:%s%s" % (self.__twenty_six(left_alpha, size), left_int,
                              self.__twenty_six(right_alpha, size), right_int)

    @staticmethod
    def __style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param cell_range:
        :param alignment:
        :param ws:  Excel worksheet instance
        :param cell_range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            t = row[0]
            r = row[-1]
            t.border = t.border + left
            r.border = r.border + right
            if fill:
                for _c in row:
                    _c.fill = fill

    def save(self):

        self.write_wb.save(self.__write_file)


def main():
    c = InsertColumn('test.xlsx', 'new.xlsx')
    c.insert_col(2)
    c.save()

    c = InsertColumn('test.xls', 'old.xls')
    c.insert_col(2)
    c.save()


if __name__ == '__main__':
    main()
